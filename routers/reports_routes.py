"""
routers/reports_routes.py — Phase 6 reporting, role-scoped.

Scope rules:
  Transcripts : HC = all | Coach = self + own clients | Client = self only
  Credits     : HC = all | Coach = received by them + allocations they made
                | Client = credits they received
  Usage       : HC = all users | Coach = self + own clients | Client = self
"""
from datetime import datetime, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session

from database import (get_db, User, Skill, Conversation, Message,
                      CreditTransaction, UsageMonthly)
from auth import get_current_user

router = APIRouter(prefix="/api/reports", tags=["reports"])


# ---------- Scope helpers ----------

def visible_user_ids(user: User, db: Session):
    """Return list of user ids this user may report on, or None for 'all' (HC)."""
    if user.role == "head_coach":
        return None
    if user.role == "coach":
        ids = [u.id for u in db.query(User.id).filter(User.coach_id == user.id,
                                                      User.role == "client")]
        return [user.id] + list(ids)
    return [user.id]


def scoped_users(user: User, db: Session):
    """Users listable in report filters."""
    if user.role == "head_coach":
        rows = db.query(User).filter(User.role != "head_coach").order_by(User.name).all()
        rows = [user] + rows
    elif user.role == "coach":
        clients = db.query(User).filter(User.coach_id == user.id,
                                        User.role == "client").order_by(User.name).all()
        rows = [user] + clients
    else:
        rows = [user]
    return rows


# ---------- Filter options ----------

@router.get("/filters")
def filter_options(user: User = Depends(get_current_user),
                   db: Session = Depends(get_db)):
    users = scoped_users(user, db)
    skills = db.query(Skill).order_by(Skill.title).all()
    return {"users": [{"id": u.id, "name": u.name, "role": u.role} for u in users],
            "skills": [{"id": s.id, "title": s.title} for s in skills]}


# ---------- Conversation transcripts ----------

@router.get("/conversations")
def report_conversations(user_id: Optional[int] = None,
                         skill_id: Optional[int] = None,
                         date_from: Optional[str] = None,   # YYYY-MM-DD
                         date_to: Optional[str] = None,
                         user: User = Depends(get_current_user),
                         db: Session = Depends(get_db)):
    allowed = visible_user_ids(user, db)
    q = (db.query(Conversation, Skill, User)
         .join(Skill, Conversation.skill_id == Skill.id)
         .join(User, Conversation.user_id == User.id))
    if allowed is not None:
        q = q.filter(Conversation.user_id.in_(allowed))
    if user_id:
        if allowed is not None and user_id not in allowed:
            raise HTTPException(403, "You cannot report on this user")
        q = q.filter(Conversation.user_id == user_id)
    if skill_id:
        q = q.filter(Conversation.skill_id == skill_id)
    if date_from:
        q = q.filter(Conversation.created_at >= datetime.fromisoformat(date_from))
    if date_to:
        q = q.filter(Conversation.created_at <
                     datetime.fromisoformat(date_to) + timedelta(days=1))
    rows = q.order_by(Conversation.id.desc()).limit(300).all()

    out = []
    for conv, skill, owner in rows:
        msg_count = db.query(Message).filter(
            Message.conversation_id == conv.id).count()
        out.append({"id": conv.id, "title": conv.title, "skill": skill.title,
                    "user": owner.name, "user_role": owner.role,
                    "model_id": conv.model_id, "messages": msg_count,
                    "created_at": str(conv.created_at)})
    return out


@router.get("/conversations/{conv_id}")
def report_transcript(conv_id: int, user: User = Depends(get_current_user),
                      db: Session = Depends(get_db)):
    conv = db.query(Conversation).get(conv_id)
    if not conv:
        raise HTTPException(404, "Conversation not found")
    allowed = visible_user_ids(user, db)
    if allowed is not None and conv.user_id not in allowed:
        raise HTTPException(403, "You cannot view this conversation")
    skill = db.query(Skill).get(conv.skill_id)
    owner = db.query(User).get(conv.user_id)
    msgs = (db.query(Message).filter(Message.conversation_id == conv.id)
            .order_by(Message.id).all())
    return {"id": conv.id, "title": conv.title, "skill": skill.title,
            "user": owner.name, "model_id": conv.model_id,
            "created_at": str(conv.created_at),
            "messages": [{"role": m.role, "content": m.content,
                          "superseded": bool(m.superseded),
                          "tokens": m.input_tokens + m.output_tokens,
                          "at": str(m.created_at)} for m in msgs]}


# ---------- Credit ledger (audit trail) ----------

@router.get("/credits")
def report_credits(user: User = Depends(get_current_user),
                   db: Session = Depends(get_db)):
    q = db.query(CreditTransaction)
    if user.role == "coach":
        client_ids = [u.id for u in db.query(User.id).filter(
            User.coach_id == user.id, User.role == "client")]
        q = q.filter(
            (CreditTransaction.to_user_id == user.id) |
            (CreditTransaction.from_user_id == user.id) |
            (CreditTransaction.to_user_id.in_(client_ids)))
    elif user.role == "client":
        q = q.filter((CreditTransaction.to_user_id == user.id) |
                     (CreditTransaction.from_user_id == user.id))
    rows = q.order_by(CreditTransaction.id.desc()).limit(500).all()

    users = {u.id: u for u in db.query(User).all()}   # fetched ONCE

    def ser(t):
        fu, tu = users.get(t.from_user_id), users.get(t.to_user_id)
        return {"id": t.id,
                "from": fu.name if fu else "Head Coach (platform)",
                "from_role": fu.role.replace("_", " ") if fu else "head coach",
                "to": tu.name if tu else "?",
                "to_role": tu.role.replace("_", " ") if tu else "?",
                "tokens": t.tokens, "note": t.note, "at": str(t.created_at)}
    return [ser(t) for t in rows]


# ---------- Usage summary ----------

@router.get("/usage")
def report_usage(month: Optional[str] = None,   # "YYYY-MM", default current
                 user: User = Depends(get_current_user),
                 db: Session = Depends(get_db)):
    month = month or datetime.utcnow().strftime("%Y-%m")
    allowed = visible_user_ids(user, db)
    q = (db.query(UsageMonthly, User)
         .join(User, UsageMonthly.user_id == User.id)
         .filter(UsageMonthly.month == month))
    if allowed is not None:
        q = q.filter(UsageMonthly.user_id.in_(allowed))
    rows = q.order_by(UsageMonthly.tokens_used.desc()).all()
    return {"month": month,
            "rows": [{"user": u.name, "role": u.role,
                      "tokens_used": um.tokens_used,
                      "balance": u.token_balance or 0}
                     for um, u in rows]}


# ---------- Exports ----------

from io import BytesIO
from fastapi.responses import StreamingResponse, PlainTextResponse


def xlsx_response(headers: list, data_rows: list, sheet: str, filename: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in data_rows:
        ws.append(r)
    for i, _ in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 22
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/export/conversations")
def export_conversations(user_id: Optional[int] = None,
                         skill_id: Optional[int] = None,
                         date_from: Optional[str] = None,
                         date_to: Optional[str] = None,
                         user: User = Depends(get_current_user),
                         db: Session = Depends(get_db)):
    rows = report_conversations(user_id, skill_id, date_from, date_to, user, db)
    return xlsx_response(
        ["ID", "Title", "User", "Role", "Skill", "Model", "Messages", "Started (UTC)"],
        [[r["id"], r["title"], r["user"], r["user_role"], r["skill"],
          r["model_id"], r["messages"], r["created_at"]] for r in rows],
        "Transcripts", "skillcoach_transcripts.xlsx")


@router.get("/export/credits")
def export_credits(user: User = Depends(get_current_user),
                   db: Session = Depends(get_db)):
    rows = report_credits(user, db)
    return xlsx_response(
        ["Credit ID", "From", "From role", "To", "To role", "Tokens", "Note", "When (UTC)"],
        [[r["id"], r["from"], r["from_role"], r["to"], r["to_role"],
          r["tokens"], r["note"], r["at"]] for r in rows],
        "Credit Ledger", "skillcoach_credit_ledger.xlsx")


@router.get("/export/usage")
def export_usage(month: Optional[str] = None,
                 user: User = Depends(get_current_user),
                 db: Session = Depends(get_db)):
    r = report_usage(month, user, db)
    return xlsx_response(
        ["User", "Role", "Tokens used", "Current balance"],
        [[x["user"], x["role"], x["tokens_used"], x["balance"]] for x in r["rows"]],
        f"Usage {r['month']}", f"skillcoach_usage_{r['month']}.xlsx")


@router.get("/conversations/{conv_id}/download")
def download_transcript(conv_id: int, user: User = Depends(get_current_user),
                        db: Session = Depends(get_db)):
    """Formatted HTML transcript. Assistant messages are Markdown-rendered;
    superseded (edited-away) turns are shown greyed with a label."""
    import html as htmlmod
    import markdown as md

    t = report_transcript(conv_id, user, db)

    blocks = []
    for m in t["messages"]:
        if m["role"] == "assistant":
            body = md.markdown(m["content"], extensions=["tables", "fenced_code"])
        else:
            body = "<p>" + htmlmod.escape(m["content"]).replace("\n", "<br>") + "</p>"
        cls = m["role"] + (" superseded" if m.get("superseded") else "")
        label = ("You" if m["role"] == "user" else "Assistant") + \
                (" · edited away" if m.get("superseded") else "")
        blocks.append(
            f'<div class="msg {cls}"><div class="who">{label} · {m["at"][:16]}</div>{body}</div>')

    page = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<title>{htmlmod.escape(t['title'])}</title>
<style>
 body {{ font-family: Arial, Helvetica, sans-serif; max-width: 760px;
        margin: 30px auto; padding: 0 16px; color: #222; }}
 h1 {{ font-size: 20px; }} .meta {{ color: #777; font-size: 13px; margin-bottom: 24px; }}
 .msg {{ border-radius: 10px; padding: 12px 16px; margin: 12px 0; }}
 .msg.user {{ background: #eef1ff; margin-left: 15%; }}
 .msg.assistant {{ background: #f5f5f5; margin-right: 15%; }}
 .msg.superseded {{ opacity: .55; border: 1px dashed #bbb; }}
 .who {{ font-size: 11px; font-weight: bold; color: #888; margin-bottom: 6px;
        text-transform: uppercase; letter-spacing: .05em; }}
 pre {{ background: #2d2d2d; color: #eee; padding: 10px; border-radius: 6px; overflow-x: auto; }}
 code {{ background: #e8e8e8; padding: 1px 4px; border-radius: 3px; }}
 pre code {{ background: none; padding: 0; }}
 table {{ border-collapse: collapse; }} td, th {{ border: 1px solid #ccc; padding: 5px 9px; }}
</style></head><body>
<h1>{htmlmod.escape(t['title'])}</h1>
<div class="meta">User: {htmlmod.escape(t['user'])} · Skill: {htmlmod.escape(t['skill'])}
 · Model: {htmlmod.escape(t['model_id'] or '')} · Started (UTC): {t['created_at'][:16]}</div>
{''.join(blocks)}
</body></html>"""

    from fastapi.responses import HTMLResponse
    return HTMLResponse(page, headers={
        "Content-Disposition": f'attachment; filename="transcript_{conv_id}.html"'})
